import os
import uuid
import csv
from datetime import datetime, timedelta, timezone
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, abort, send_file, jsonify
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect
from PIL import Image
from werkzeug.security import generate_password_hash, check_password_hash

# Import database models and settings
from config import Config
from models import db, User, CompanySettings, Invoice, InvoiceItem, Customer, ActivityLog
from pdf_generator import generate_invoice_pdf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 1. Initialize Application & Middleware Extensions
app = Flask(__name__)
app.config.from_object(Config)

# Hardened session cookies
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV', 'development') == 'production'

# FIX: Bind SQLAlchemy to Flask app at module level so it works under
# both 'python app.py' and 'flask run' / gunicorn deployments.
db.init_app(app)

# Enable password hashing
bcrypt = Bcrypt(app)

# Enable user session management
login_manager = LoginManager(app)
login_manager.login_view = 'login' # Redirect route for unauthenticated sessions
login_manager.login_message_category = 'warning' # CSS alert color classification

# Enable global CSRF protection on forms
csrf = CSRFProtect(app)

# Secure PDF Storage Location (Outside public static path)
SECURE_PDF_FOLDER = os.path.join(app.config['BASE_DIR'], 'generated_invoices')

# Exports directory for downloadable report files
EXPORTS_FOLDER = os.path.join(app.config['BASE_DIR'], 'exports')

# Ensure necessary system folders exist on boot
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(SECURE_PDF_FOLDER, exist_ok=True)
os.makedirs(EXPORTS_FOLDER, exist_ok=True)


# 2. Authentication Helper: User Loader
@login_manager.user_loader
def load_user(user_id):
    """Flask-Login callback to retrieve user details from SQLite session keys."""
    return db.session.get(User, int(user_id))


# 3. Secure Sequential Invoice Number Generator
def get_next_invoice_number():
    """
    Generates a secure, sequential invoice number.
    Format: INV-YYYY-NNNN (e.g. INV-2026-0001)
    """
    year = datetime.now(timezone.utc).year
    # Fetch the last created invoice to determine sequence ID
    last_invoice = Invoice.query.order_by(Invoice.id.desc()).first()
    
    if last_invoice:
        next_sequence = last_invoice.id + 1
    else:
        next_sequence = 1
        
    return f"INV-{year}-{next_sequence:04d}"


# Helper to validate image upload extensions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


# 4. Routing Controllers

# AUTHENTICATION: Check Email (AJAX endpoint)
@app.route('/api/check-email', methods=['POST'])
@csrf.exempt
def check_email():
    data = request.get_json() or {}
    email = data.get('email', '').strip()
    exists = User.query.filter_by(email=email).first() is not None
    return jsonify({'exists': exists})


# AUTHENTICATION: Login
@app.route('/login', methods=['GET', 'POST'])
def login():
    # Redirect if already authenticated
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        remember = True if request.form.get('remember') else False
        
        user = User.query.filter_by(email=email).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user, remember=remember)
            flash("Welcome back! Logged in successfully.", "success")
            return redirect(url_for('dashboard'))
        else:
            flash("Invalid email or password.", "error")
            
    return render_template('login.html')


# AUTHENTICATION: Register
@app.route('/register', methods=['POST'])
def register():
    # Redirect if already authenticated
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    full_name = request.form.get('full_name', '').strip()
    company_name = request.form.get('company_name', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    password = request.form.get('password', '')
    confirm_password = request.form.get('confirm_password', '')
    
    # Validation
    if not full_name:
        flash("Full Name is required.", "error_full_name")
    if not company_name:
        flash("Company Name is required.", "error_company_name")
    if not email:
        flash("Email is required.", "error_email")
    if not password:
        flash("Password is required.", "error_password")
        
    if not (full_name and company_name and email and password):
        return redirect(url_for('login', tab='signup'))
        
    # Email uniqueness
    if User.query.filter_by(email=email).first() is not None:
        flash("Email address already registered.", "error_email")
        return redirect(url_for('login', tab='signup'))
        
    # Password complexity validation: min 8 chars, 1 uppercase, 1 lowercase, 1 digit
    if (len(password) < 8 or
        not any(c.isupper() for c in password) or
        not any(c.islower() for c in password) or
        not any(c.isdigit() for c in password)):
        flash("Password must be at least 8 characters long and contain at least one uppercase letter, one lowercase letter, and one number.", "error_password")
        return redirect(url_for('login', tab='signup'))
        
    if password != confirm_password:
        flash("Passwords do not match.", "error_confirm_password")
        return redirect(url_for('login', tab='signup'))
        
    # Creation
    hashed_pw = generate_password_hash(password)
    new_user = User(
        full_name=full_name,
        company_name=company_name,
        email=email,
        phone=phone if phone else None,
        password_hash=hashed_pw
    )
    db.session.add(new_user)
    db.session.commit()
    
    login_user(new_user)
    flash("Account created and logged in successfully!", "success")
    return redirect(url_for('dashboard'))


# AUTHENTICATION: Forgot Password
@app.route('/forgot-password')
def forgot_password():
    return render_template('forgot_password.html')


# AUTHENTICATION: User Profile
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        company_name = request.form.get('company_name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        current_password = request.form.get('current_password', '')
        
        # Validation
        if not full_name:
            flash("Full Name is required.", "error")
            return redirect(url_for('profile'))
        if not company_name:
            flash("Company Name is required.", "error")
            return redirect(url_for('profile'))
        if not email:
            flash("Email is required.", "error")
            return redirect(url_for('profile'))
            
        # Email change requires current password validation
        email_changed = (email.lower() != current_user.email.lower())
        if email_changed:
            if not current_password:
                flash("Current password is required to change email.", "error")
                return redirect(url_for('profile'))
            if not check_password_hash(current_user.password_hash, current_password):
                flash("Incorrect current password.", "error")
                return redirect(url_for('profile'))
            if User.query.filter_by(email=email).first() is not None:
                flash("Email address is already in use.", "error")
                return redirect(url_for('profile'))
                
        current_user.full_name = full_name
        current_user.company_name = company_name
        current_user.phone = phone if phone else None
        if email_changed:
            current_user.email = email
            
        current_user.updated_at = datetime.utcnow()
        db.session.commit()
        flash("Profile updated successfully.", "success")
        return redirect(url_for('profile'))
        
    return render_template('profile.html')


# AUTHENTICATION: Logout
@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash("You have been logged out successfully.", "success")
    return redirect(url_for('dashboard'))


# DASHBOARD: Overview & Invoice list
@app.route('/')
def dashboard():
    # If the user is not authenticated, show the public landing page:
    if not current_user.is_authenticated:
        return render_template('index.html')
        
    # UX Check: Redirect to settings if company is not configured yet
    company = CompanySettings.query.first()
    if not company:
        flash("Welcome to BillFlow! Please configure your company details first to start generating invoices.", "warning")
        return redirect(url_for('settings'))
        
    invoices = Invoice.query.order_by(Invoice.date_created.desc(), Invoice.id.desc()).all()
    
    # Calculate aggregate KPI statistics
    total_invoiced = sum(inv.total_amount for inv in invoices)
    total_paid = sum(inv.total_amount for inv in invoices if inv.status == 'Paid')
    total_unpaid = sum(inv.total_amount for inv in invoices if inv.status == 'Unpaid')
    
    stats = {
        'total_invoiced': total_invoiced,
        'total_paid': total_paid,
        'total_unpaid': total_unpaid,
        'count': len(invoices)
    }
    
    return render_template('dashboard.html', invoices=invoices, stats=stats, company=company)


# REPORTS: Main Dashboard (Version 4)
@app.route('/reports')
@login_required
def reports_dashboard():
    return render_template('reports/dashboard.html')



# Helper: get current time in IST (India Standard Time: UTC + 5:30)
def get_ist_now():
    from datetime import timezone
    return datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)


# Helper: calculate date range boundaries in IST
def get_date_filter_bounds(filter_type, custom_start_str=None, custom_end_str=None):
    today = get_ist_now().date()
    
    if filter_type == 'today':
        return today, today
    elif filter_type == 'yesterday':
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday
    elif filter_type == 'last_7_days':
        return today - timedelta(days=6), today
    elif filter_type == 'last_30_days':
        return today - timedelta(days=29), today
    elif filter_type == 'this_month':
        return today.replace(day=1), today
    elif filter_type == 'this_quarter':
        quarter_month = ((today.month - 1) // 3) * 3 + 1
        return today.replace(month=quarter_month, day=1), today
    elif filter_type == 'this_year':
        return today.replace(month=1, day=1), today
    elif filter_type == 'custom':
        if not custom_start_str or not custom_end_str:
            return today, today
        start_date = datetime.strptime(custom_start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end_str, '%Y-%m-%d').date()
        return start_date, end_date
    return today - timedelta(days=6), today


# Helper: parse report date range filters from request args (Version 4 Refactor)
def parse_report_dates():
    """
    Parses range parameters, start_date, and end_date from request parameters.
    Returns (start_date, end_date) or raises ValueError/exception.
    """
    filter_type = request.args.get('range', 'last_7_days').strip().lower()
    start_str = request.args.get('start_date', '').strip()
    end_str = request.args.get('end_date', '').strip()
    
    start_date, end_date = get_date_filter_bounds(filter_type, start_str, end_str)
    if start_date > end_date:
        raise ValueError("Start date cannot be after end date.")
    return start_date, end_date


# API: Summary cards statistics (Version 4 Phase 1 Optimized)
@app.route('/api/reports/summary')
@login_required
def api_reports_summary():
    try:
        start_date, end_date = parse_report_dates()
    except Exception as e:
        return {'success': False, 'message': str(e)}, 400

    # Single consolidated aggregate query for maximum performance
    stats = db.session.query(
        db.func.sum(db.case(
            (Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue']), Invoice.total_amount),
            else_=0.0
        )).label('revenue'),
        db.func.sum(db.case(
            (Invoice.status.in_(['Sent', 'Pending', 'Overdue', 'Unpaid']), Invoice.total_amount),
            else_=0.0
        )).label('outstanding'),
        db.func.count(Invoice.id).label('total_invoices'),
        db.func.count(db.distinct(Invoice.client_name)).label('total_customers'),
        db.func.sum(db.case(
            (Invoice.status == 'Paid', 1),
            else_=0
        )).label('paid_count'),
        db.func.sum(db.case(
            (Invoice.status.in_(['Pending', 'Sent']), 1),
            else_=0
        )).label('pending_count'),
        db.func.sum(db.case(
            (Invoice.status == 'Overdue', 1),
            else_=0
        )).label('overdue_count'),
        db.func.sum(db.case(
            (Invoice.status == 'Cancelled', 1),
            else_=0
        )).label('cancelled_count')
    ).filter(
        Invoice.is_deleted == False,
        Invoice.date_created.between(start_date, end_date)
    ).first()

    return {
        'success': True,
        'total_revenue': float(stats.revenue or 0.0),
        'outstanding_amount': float(stats.outstanding or 0.0),
        'total_invoices': stats.total_invoices or 0,
        'total_customers': stats.total_customers or 0,
        'paid_count': int(stats.paid_count or 0),
        'pending_count': int(stats.pending_count or 0),
        'overdue_count': int(stats.overdue_count or 0),
        'cancelled_count': int(stats.cancelled_count or 0)
    }


# API: Revenue trend database (Version 4 Phase 1)
@app.route('/api/reports/revenue')
@login_required
def api_reports_revenue():
    try:
        start_date, end_date = parse_report_dates()
    except Exception as e:
        return {'success': False, 'message': str(e)}, 400

    results = db.session.query(
        Invoice.date_created,
        db.func.sum(Invoice.total_amount)
    ).filter(
        Invoice.is_deleted == False,
        Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue']),
        Invoice.date_created.between(start_date, end_date)
    ).group_by(Invoice.date_created).order_by(Invoice.date_created.asc()).all()

    data = [{'date': str(d), 'revenue': float(r)} for d, r in results]
    return {'success': True, 'data': data}


# API: Invoice status distribution (Version 4 Phase 1)
@app.route('/api/reports/invoice-status')
@login_required
def api_reports_invoice_status():
    try:
        start_date, end_date = parse_report_dates()
    except Exception as e:
        return {'success': False, 'message': str(e)}, 400

    results = db.session.query(
        Invoice.status,
        db.func.count(Invoice.id)
    ).filter(
        Invoice.is_deleted == False,
        Invoice.date_created.between(start_date, end_date)
    ).group_by(Invoice.status).all()

    data = {status: count for status, count in results}
    return {'success': True, 'data': data}


# API: Customer analytics aggregates (Version 4 Phase 2)
@app.route('/api/reports/customers')
@login_required
def api_reports_customers():
    try:
        start_date, end_date = parse_report_dates()
    except Exception as e:
        return {'success': False, 'message': str(e)}, 400

    # Group by customer and compute aggregates
    results = db.session.query(
        Customer.id,
        Customer.name,
        Customer.email,
        db.func.sum(db.case(
            (Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue']), Invoice.total_amount),
            else_=0.0
        )).label('revenue'),
        db.func.count(Invoice.id).label('invoice_count'),
        db.func.sum(db.case(
            (Invoice.status.in_(['Sent', 'Pending', 'Overdue', 'Unpaid']), Invoice.total_amount),
            else_=0.0
        )).label('outstanding'),
        db.func.avg(Invoice.total_amount).label('avg_value'),
        db.func.max(Invoice.date_created).label('last_date')
    ).select_from(Customer)\
     .outerjoin(Invoice, (Invoice.customer_id == Customer.id) & (Invoice.is_deleted == False) & (Invoice.date_created.between(start_date, end_date)))\
     .group_by(Customer.id, Customer.name, Customer.email)\
     .order_by(db.desc('revenue')).all()

    data = []
    for r in results:
        data.append({
            'id': r.id,
            'name': r.name,
            'email': r.email,
            'revenue': float(r.revenue or 0.0),
            'invoice_count': r.invoice_count or 0,
            'outstanding': float(r.outstanding or 0.0),
            'avg_value': float(r.avg_value or 0.0),
            'last_date': r.last_date.strftime('%Y-%m-%d') if r.last_date else '-'
        })

    return {'success': True, 'data': data}


# REPORTS: Customer profile details (Version 4 Phase 2)
@app.route('/reports/customers/<int:customer_id>')
@login_required
def reports_customer_profile(customer_id):
    customer = db.session.get(Customer, customer_id)
    if not customer:
        abort(404)

    # Compute lifetime stats
    revenue = db.session.query(db.func.sum(Invoice.total_amount)).filter(
        Invoice.customer_id == customer_id,
        Invoice.is_deleted == False,
        Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue'])
    ).scalar() or 0.0

    outstanding = db.session.query(db.func.sum(Invoice.total_amount)).filter(
        Invoice.customer_id == customer_id,
        Invoice.is_deleted == False,
        Invoice.status.in_(['Sent', 'Pending', 'Overdue', 'Unpaid'])
    ).scalar() or 0.0

    # Retrieve invoices
    invoices = Invoice.query.filter_by(customer_id=customer_id, is_deleted=False).order_by(Invoice.date_created.desc()).all()

    # Revenue trend grouped by month (SQLite/PostgreSQL agnostic calculation in Python)
    trend_raw = db.session.query(Invoice.date_created, Invoice.total_amount).filter(
        Invoice.customer_id == customer_id,
        Invoice.is_deleted == False,
        Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue'])
    ).order_by(Invoice.date_created.asc()).all()

    from collections import defaultdict
    monthly_trend = defaultdict(float)
    for date_val, amount in trend_raw:
        month_str = date_val.strftime('%Y-%m')
        monthly_trend[month_str] += amount

    sorted_trend = sorted(monthly_trend.items())
    trend_labels = [item[0] for item in sorted_trend]
    trend_values = [item[1] for item in sorted_trend]

    return render_template(
        'reports/customer_profile.html',
        customer=customer,
        revenue=revenue,
        outstanding=outstanding,
        invoices=invoices,
        trend_labels=trend_labels,
        trend_values=trend_values
    )



# API: Product analytics (Version 4 Phase 3)
@app.route('/api/reports/products')
@login_required
def api_reports_products():
    try:
        start_date, end_date = parse_report_dates()
    except Exception as e:
        return {'success': False, 'message': str(e)}, 400

    # Query all non-deleted, non-draft, non-cancelled invoice items within period
    items = db.session.query(
        InvoiceItem.description,
        InvoiceItem.quantity,
        InvoiceItem.unit_price,
        InvoiceItem.hsn_sac,
        InvoiceItem.total
    ).join(Invoice).filter(
        Invoice.is_deleted == False,
        Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue']),
        Invoice.date_created.between(start_date, end_date)
    ).all()

    # Aggregate in Python for maximum flexibility, database compatibility, and description normalization
    product_stats = {}
    
    service_keywords = ['service', 'consulting', 'support', 'fee', 'maintenance', 'installation', 'labor', 'work', 'hours', 'rent', 'training']

    for item in items:
        desc_norm = item.description.strip().lower() if item.description else ''
        if not desc_norm:
            continue
            
        hsn = item.hsn_sac.strip() if item.hsn_sac else ""
        
        # Classification
        is_service = False
        if hsn.startswith('99'):
            is_service = True
        else:
            if any(kw in desc_norm for kw in service_keywords):
                is_service = True
                
        if desc_norm not in product_stats:
            product_stats[desc_norm] = {
                'name': item.description.strip(),
                'quantity': 0,
                'revenue': 0.0,
                'is_service': is_service
            }
        product_stats[desc_norm]['quantity'] += item.quantity
        product_stats[desc_norm]['revenue'] += item.total

    # Compile lists
    all_aggregated = []
    for key, stats in product_stats.items():
        avg_price = stats['revenue'] / stats['quantity'] if stats['quantity'] > 0 else 0.0
        all_aggregated.append({
            'name': stats['name'],
            'quantity': stats['quantity'],
            'revenue': stats['revenue'],
            'avg_price': avg_price,
            'is_service': stats['is_service']
        })

    # Sort lists
    products_only = [x for x in all_aggregated if not x['is_service']]
    services_only = [x for x in all_aggregated if x['is_service']]

    # Top 10 Products by Quantity
    top_products_qty = sorted(products_only, key=lambda x: x['quantity'], reverse=True)[:10]
    # Top 10 Services by Quantity
    top_services_qty = sorted(services_only, key=lambda x: x['quantity'], reverse=True)[:10]
    # Highest Revenue Products
    highest_revenue_products = sorted(products_only, key=lambda x: x['revenue'], reverse=True)[:10]

    return {
        'success': True,
        'top_products': top_products_qty,
        'top_services': top_services_qty,
        'highest_revenue_products': highest_revenue_products
    }


# API: GST Reports (Version 4 Phase 4)
@app.route('/api/reports/gst')
@login_required
def api_reports_gst():
    try:
        start_date, end_date = parse_report_dates()
    except Exception as e:
        return {'success': False, 'message': str(e)}, 400

    # Query all active invoices in period
    invoices = Invoice.query.filter(
        Invoice.is_deleted == False,
        Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue']),
        Invoice.date_created.between(start_date, end_date)
    ).order_by(Invoice.date_created.asc()).all()

    cgst_total = 0.0
    sgst_total = 0.0
    igst_total = 0.0
    gst_total = 0.0
    taxable_revenue_total = 0.0

    invoices_data = []

    for inv in invoices:
        discount_amount = inv.subtotal * (inv.discount / 100.0)
        taxable_value = inv.subtotal - discount_amount
        
        cgst_total += inv.cgst or 0.0
        sgst_total += inv.sgst or 0.0
        igst_total += inv.igst or 0.0
        gst_total += inv.tax_amount or 0.0
        taxable_revenue_total += taxable_value

        invoices_data.append({
            'invoice_number': inv.invoice_number,
            'date': inv.date_created.strftime('%Y-%m-%d'),
            'client_name': inv.client_name,
            'client_gstin': inv.client_gstin or '-',
            'taxable_value': float(taxable_value),
            'cgst': float(inv.cgst or 0.0),
            'sgst': float(inv.sgst or 0.0),
            'igst': float(inv.igst or 0.0),
            'tax_amount': float(inv.tax_amount or 0.0),
            'total_amount': float(inv.total_amount or 0.0)
        })

    return {
        'success': True,
        'cgst_collected': float(cgst_total),
        'sgst_collected': float(sgst_total),
        'igst_collected': float(igst_total),
        'total_gst': float(gst_total),
        'taxable_revenue': float(taxable_revenue_total),
        'invoices': invoices_data
    }


def get_filtered_invoices_query():
    """
    Shared query helper for retrieving invoices according to B2B parameters:
    search, status, customer_id, from_date, to_date, favorites_only.
    Enforces that deleted invoices are excluded.
    """
    query = Invoice.query.filter_by(is_deleted=False)
    
    # Customer ID filter
    customer_id = request.args.get('customer_id', type=int)
    if customer_id:
        query = query.filter(Invoice.customer_id == customer_id)
        
    # Search filter (Invoice number or client name)
    search_query = request.args.get('search', '').strip()
    if search_query:
        query = query.filter(
            (Invoice.invoice_number.ilike(f"%{search_query}%")) |
            (Invoice.client_name.ilike(f"%{search_query}%"))
        )
        
    # Status filter
    status_filter = request.args.get('status', '').strip()
    if status_filter:
        query = query.filter(Invoice.status == status_filter)
        
    # Date Range filter
    from_date_str = request.args.get('from_date', '').strip()
    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            query = query.filter(Invoice.date_created >= from_date)
        except ValueError:
            pass
            
    to_date_str = request.args.get('to_date', '').strip()
    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
            query = query.filter(Invoice.date_created <= to_date)
        except ValueError:
            pass

    # Favorites filter
    favorites_only = request.args.get('favorites_only', type=int)
    if favorites_only == 1:
        query = query.filter(Invoice.is_favorite == True)
        
    return query


# INVOICE: List Invoices (Version 3)
@app.route('/invoices')
@login_required
def list_invoices():
    query = get_filtered_invoices_query()
    
    export_format = request.args.get('export', '').strip().lower()
    if export_format in ['excel', 'csv']:
        # Fetch the exact same filtered invoices
        invoices = query.order_by(Invoice.is_favorite.desc(), Invoice.created_at.desc()).all()
        
        # Ensure exports directory exists
        exports_dir = os.path.join(app.root_path, 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        
        if export_format == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoices Export"
            
            headers = ["Invoice Number", "Customer Name", "Customer Email", "Invoice Date", "Due Date", "Subtotal", "Tax Amount", "Total Amount", "Status", "Starred"]
            ws.append(headers)
            
            from openpyxl.styles import Font, PatternFill
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                
            for inv in invoices:
                ws.append([
                    inv.invoice_number,
                    inv.client_name,
                    inv.client_email,
                    inv.date_created.strftime('%Y-%m-%d'),
                    inv.due_date.strftime('%Y-%m-%d'),
                    inv.subtotal,
                    inv.tax_amount,
                    inv.total_amount,
                    inv.status,
                    "Yes" if inv.is_favorite else "No"
                ])
                
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            export_filename = f"invoices_export_{timestamp}.xlsx"
            export_path = os.path.join(exports_dir, export_filename)
            wb.save(export_path)
            wb.close()
            
            log = ActivityLog(
                invoice_id=None,
                invoice_number="EXPORTS",
                action="Invoices Exported",
                details=f"Exported {len(invoices)} invoices as Excel sheet: '{export_filename}'."
            )
            db.session.add(log)
            db.session.commit()
            
            return send_file(
                export_path,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=export_filename
            )
            
        elif export_format == 'csv':
            import csv
            export_filename = f"invoices_export_{timestamp}.csv"
            export_path = os.path.join(exports_dir, export_filename)
            
            with open(export_path, mode='w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Invoice Number", "Customer Name", "Customer Email", "Invoice Date", "Due Date", "Subtotal", "Tax Amount", "Total Amount", "Status", "Starred"])
                for inv in invoices:
                    writer.writerow([
                        inv.invoice_number,
                        inv.client_name,
                        inv.client_email,
                        inv.date_created.strftime('%Y-%m-%d'),
                        inv.due_date.strftime('%Y-%m-%d'),
                        inv.subtotal,
                        inv.tax_amount,
                        inv.total_amount,
                        inv.status,
                        "Yes" if inv.is_favorite else "No"
                    ])
                    
            log = ActivityLog(
                invoice_id=None,
                invoice_number="EXPORTS",
                action="Invoices Exported",
                details=f"Exported {len(invoices)} invoices as CSV: '{export_filename}'."
            )
            db.session.add(log)
            db.session.commit()
            
            return send_file(
                export_path,
                mimetype="text/csv",
                as_attachment=True,
                download_name=export_filename
            )
            
    # Favorites pinned first, then newest invoices first
    invoices = query.order_by(Invoice.is_favorite.desc(), Invoice.created_at.desc()).all()
    customers = Customer.query.order_by(Customer.name).all()
    today_date = datetime.now(timezone.utc).date()
    
    return render_template(
        'invoices/list.html',
        invoices=invoices,
        customers=customers,
        today_date=today_date
    )


# INVOICE: Trash Bin (Version 3)
@app.route('/invoices/trash')
@login_required
def trash_invoices():
    invoices = Invoice.query.filter_by(is_deleted=True).order_by(Invoice.deleted_at.desc()).all()
    return render_template('invoices/trash.html', invoices=invoices)


# INVOICE: Toggle Favorite Pin (Version 3)
@app.route('/invoices/<int:invoice_id>/favorite', methods=['POST'])
@login_required
def toggle_invoice_favorite(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
    invoice.is_favorite = not invoice.is_favorite
    
    # Log activity
    state = "Starred" if invoice.is_favorite else "Unstarred"
    log = ActivityLog(
        invoice_id=invoice.id,
        invoice_number=invoice.invoice_number,
        action=f"Invoice {state}",
        details=f"Invoice marked as {state.lower()}."
    )
    db.session.add(log)
    db.session.commit()
    
    # Redirect back preserving filters
    search = request.form.get('search', '')
    status = request.form.get('status', '')
    customer_id = request.form.get('customer_id', '')
    from_date = request.form.get('from_date', '')
    to_date = request.form.get('to_date', '')
    favorites_only = request.form.get('favorites_only', '')
    
    return redirect(url_for(
        'list_invoices',
        search=search,
        status=status,
        customer_id=customer_id,
        from_date=from_date,
        to_date=to_date,
        favorites_only=favorites_only
    ))


# INVOICE: View Invoice Details (Version 3)
@app.route('/invoices/<int:invoice_id>')
@login_required
def view_invoice_details(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
    company = CompanySettings.query.first()
    logs = ActivityLog.query.filter_by(invoice_number=invoice.invoice_number).order_by(ActivityLog.created_at.desc()).all()
    today_date = datetime.now(timezone.utc).date()
    
    # Calculate HSN groups if present
    hsn_groups = {}
    has_hsn = False
    for item in invoice.items:
        if item.hsn_sac:
            has_hsn = True
            
    if has_hsn:
        for item in invoice.items:
            code = item.hsn_sac.strip() if item.hsn_sac else "-"
            if code not in hsn_groups:
                hsn_groups[code] = {
                    'taxable_value': 0.0,
                    'cgst': 0.0,
                    'sgst': 0.0,
                    'igst': 0.0,
                    'total_tax': 0.0
                }
            discount_fraction = invoice.discount / 100.0
            taxable_val = item.total * (1.0 - discount_fraction)
            tax_amount = taxable_val * (item.tax_rate / 100.0)
            
            if invoice.cgst > 0:
                cgst = tax_amount / 2.0
                sgst = tax_amount / 2.0
                igst = 0.0
            else:
                cgst = 0.0
                sgst = 0.0
                igst = tax_amount
                
            hsn_groups[code]['taxable_value'] += taxable_val
            hsn_groups[code]['cgst'] += cgst
            hsn_groups[code]['sgst'] += sgst
            hsn_groups[code]['igst'] += igst
            hsn_groups[code]['total_tax'] += tax_amount

    return render_template(
        'invoices/details.html',
        invoice=invoice,
        company=company,
        logs=logs,
        today_date=today_date,
        has_hsn=has_hsn,
        hsn_groups=hsn_groups
    )


# INVOICE: Update Status (Version 3)
@app.route('/invoices/<int:invoice_id>/status', methods=['POST'])
@login_required
def update_invoice_status(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
        
    old_status = invoice.status
    new_status = request.form.get('status', '').strip()
    if new_status not in ['Draft', 'Sent', 'Paid', 'Pending', 'Overdue', 'Cancelled']:
        flash("Invalid status value.", "error")
        return redirect(url_for('view_invoice_details', invoice_id=invoice.id))
        
    invoice.status = new_status
    
    # Log activity
    log = ActivityLog(
        invoice_id=invoice.id,
        invoice_number=invoice.invoice_number,
        action="Status Changed",
        details=f"Status changed from '{old_status}' to '{new_status}'."
    )
    db.session.add(log)
    db.session.commit()
    
    flash(f"Invoice status updated to {new_status} successfully.", "success")
    return redirect(url_for('view_invoice_details', invoice_id=invoice.id))


# INVOICE: Move to Trash (Soft Delete) (Version 3)
@app.route('/invoices/<int:invoice_id>/trash', methods=['POST'])
@login_required
def trash_invoice(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
    invoice.is_deleted = True
    invoice.deleted_at = datetime.now(timezone.utc)
    
    # Log activity
    log = ActivityLog(
        invoice_id=invoice.id,
        invoice_number=invoice.invoice_number,
        action="Invoice Trashed",
        details=f"Invoice moved to Trash Bin."
    )
    db.session.add(log)
    db.session.commit()
    
    flash(f"Invoice '{invoice.invoice_number}' moved to Trash.", "success")
    return redirect(url_for('list_invoices'))


# INVOICE: Restore from Trash (Version 3)
@app.route('/invoices/<int:invoice_id>/restore', methods=['POST'])
@login_required
def restore_invoice(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
    invoice.is_deleted = False
    invoice.deleted_at = None
    
    # Log activity
    log = ActivityLog(
        invoice_id=invoice.id,
        invoice_number=invoice.invoice_number,
        action="Invoice Restored",
        details=f"Invoice restored from Trash Bin."
    )
    db.session.add(log)
    db.session.commit()
    
    flash(f"Invoice '{invoice.invoice_number}' restored successfully.", "success")
    return redirect(url_for('trash_invoices'))


# INVOICE: Permanent Delete (Version 3)
@app.route('/invoices/<int:invoice_id>/delete-permanent', methods=['POST'])
@login_required
def delete_invoice_permanent(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
        
    num = invoice.invoice_number
    try:
        # Delete compiled PDF if exists on disk
        pdf_filename = f"{num}.pdf"
        pdf_path = os.path.join(SECURE_PDF_FOLDER, pdf_filename)
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
            
        # Log activity BEFORE deletion from DB (sets invoice_id to None but keeps number)
        log = ActivityLog(
            invoice_id=None,
            invoice_number=num,
            action="Invoice Deleted Permanently",
            details=f"Invoice was permanently deleted and PDF removed from disk."
        )
        db.session.add(log)
        db.session.delete(invoice)
        db.session.commit()
        flash(f"Invoice '{num}' permanently deleted.", "success")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Permanent delete failed: {e}", exc_info=True)
        flash(f"Failed to permanently delete invoice: {e}", "error")
        
    return redirect(url_for('trash_invoices'))


# INVOICE: Activity Log (Version 3)
@app.route('/invoices/activity-log')
@login_required
def activity_logs():
    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).all()
    return render_template('invoices/activity_log.html', logs=logs)


# INVOICE: Create New Invoice
@app.route('/invoices/new', methods=['GET', 'POST'])
@login_required
def create_invoice():
    company = CompanySettings.query.first()
    if not company:
        flash("You must complete company settings before generating invoices.", "warning")
        return redirect(url_for('settings'))
        
    if request.method == 'POST':
        # Retrieve form data
        client_name = request.form.get('client_name', '').strip()
        client_email = request.form.get('client_email', '').strip()
        client_phone = request.form.get('client_phone', '').strip()
        client_address = request.form.get('client_address', '').strip()
        client_gstin = request.form.get('client_gstin', '').strip().upper()
        
        status = request.form.get('status', 'Unpaid')
        notes = request.form.get('notes', '').strip()
        
        # Validation checks
        if not client_name or not client_email or not client_address:
            flash("Client name, email, and address are required.", "error")
            return redirect(url_for('create_invoice'))
            
        # Parse adjustments
        try:
            gst_rate = float(request.form.get('gst_rate', 18.0))
            discount = float(request.form.get('discount', 0.0))
            if discount < 0 or discount > 100:
                discount = 0.0
        except ValueError:
            gst_rate = 18.0
            discount = 0.0

        # Retrieve itemized lists
        descriptions = request.form.getlist('description[]')
        hsn_sacs = request.form.getlist('hsn_sac[]')
        quantities = request.form.getlist('quantity[]')
        unit_prices = request.form.getlist('unit_price[]')
        
        if not descriptions or len(descriptions) == 0:
            flash("An invoice must contain at least one line item.", "error")
            return redirect(url_for('create_invoice'))

        # Calculate totals on the server (Prevents form tampering)
        subtotal = 0.0
        invoice_items = []
        
        try:
            for i in range(len(descriptions)):
                desc = descriptions[i].strip()
                hsn = hsn_sacs[i].strip() if (hsn_sacs and i < len(hsn_sacs) and hsn_sacs[i].strip()) else None
                qty = int(quantities[i])
                price = float(unit_prices[i])
                
                # Check for negative bounds
                if qty < 1 or price < 0:
                    raise ValueError("Quantity must be >= 1 and Unit Price >= 0.")
                    
                item_total = qty * price
                subtotal += item_total
                
                # Build Item instance
                item = InvoiceItem(
                    description=desc,
                    quantity=qty,
                    unit_price=price,
                    tax_rate=gst_rate,
                    hsn_sac=hsn,
                    total=item_total
                )
                invoice_items.append(item)
        except (ValueError, IndexError) as e:
            flash(f"Invalid item data: {str(e)}", "error")
            return redirect(url_for('create_invoice'))

        # Final math computations
        discount_amount = subtotal * (discount / 100)
        taxable_amount = subtotal - discount_amount
        tax_amount = taxable_amount * (gst_rate / 100)
        total_amount = taxable_amount + tax_amount

        # Database Transaction Block
        try:
            # Generate the unique sequential invoice code
            invoice_number = get_next_invoice_number()
            
            # Form Dates
            date_created_str = request.form.get('date_created')
            due_date_str = request.form.get('due_date')
            
            date_created = datetime.strptime(date_created_str, '%Y-%m-%d').date() if date_created_str else datetime.now(timezone.utc).date()
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else (datetime.now(timezone.utc) + timedelta(days=15)).date()

            # Calculate CGST, SGST, IGST based on State Codes (first 2 digits of GSTIN)
            company_state = company.gstin[:2] if (company and company.gstin and len(company.gstin) >= 2) else None
            client_state = client_gstin[:2] if (client_gstin and len(client_gstin) >= 2) else None
            
            if company_state and client_state and company_state == client_state:
                cgst = tax_amount / 2.0
                sgst = tax_amount / 2.0
                igst = 0.0
            else:
                cgst = 0.0
                sgst = 0.0
                igst = tax_amount

            # Create Invoice Object
            invoice = Invoice(
                invoice_number=invoice_number,
                date_created=date_created,
                due_date=due_date,
                status=status,
                client_name=client_name,
                client_email=client_email,
                client_phone=client_phone,
                client_address=client_address,
                client_gstin=client_gstin if client_gstin else None,
                gst_rate=gst_rate,
                discount=discount,
                subtotal=subtotal,
                tax_amount=tax_amount,
                cgst=cgst,
                sgst=sgst,
                igst=igst,
                total_amount=total_amount,
                notes=notes if notes else None
            )
            
            # Associate customer_id if exists in database
            existing_customer = Customer.query.filter_by(name=client_name).first()
            if existing_customer:
                invoice.customer_id = existing_customer.id
            
            # Map relationship items
            for item in invoice_items:
                invoice.items.append(item)
                
            db.session.add(invoice)
            db.session.commit() # Save details to DB first

            # Create ActivityLog
            log = ActivityLog(
                invoice_id=invoice.id,
                invoice_number=invoice.invoice_number,
                action="Invoice Created",
                details=f"Invoice created for client '{client_name}' with total amount of ₹{total_amount:.2f}."
            )
            db.session.add(log)
            db.session.commit()

            # Compile PDF Invoice file
            pdf_filename = f"{invoice.invoice_number}.pdf"
            pdf_path = os.path.join(SECURE_PDF_FOLDER, pdf_filename)
            
            # Generate the ReportLab PDF
            generate_invoice_pdf(invoice, company, pdf_path)
            
            flash(f"Invoice {invoice.invoice_number} created and compiled successfully!", "success")
            return redirect(url_for('list_invoices'))
            
        except Exception as e:
            db.session.rollback() # Rollback DB state on error
            app.logger.error(f"Invoice creation failed: {e}")
            flash("Failed to generate invoice. Please contact administration.", "error")
            return redirect(url_for('create_invoice'))

    # GET route: Prefill calculations parameters
    next_invoice_number = get_next_invoice_number()
    today_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    default_due_date = (datetime.now(timezone.utc) + timedelta(days=15)).strftime('%Y-%m-%d')
    
    customers = Customer.query.order_by(Customer.name).all()
    
    return render_template(
        'create_invoice.html', 
        next_invoice_number=next_invoice_number, 
        today_date=today_date, 
        default_due_date=default_due_date,
        company=company,
        customers=customers
    )


# INVOICE: Download Secure PDF
@app.route('/invoices/<int:invoice_id>/download')
@login_required
def download_invoice(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
        
    company = CompanySettings.query.first()
    pdf_filename = f"{invoice.invoice_number}.pdf"
    pdf_path = os.path.join(SECURE_PDF_FOLDER, pdf_filename)
    
    # Robust Fallback: If PDF is missing on disk, regenerate it automatically
    if not os.path.exists(pdf_path):
        if not company:
            flash("Cannot regenerate PDF: Company settings are missing.", "error")
            return redirect(url_for('dashboard'))
        try:
            generate_invoice_pdf(invoice, company, pdf_path)
        except Exception as e:
            app.logger.error(f"PDF regeneration failed: {e}")
            flash("Failed to generate PDF invoice file.", "error")
            return redirect(url_for('dashboard'))
            
    return send_from_directory(
        SECURE_PDF_FOLDER, 
        pdf_filename, 
        as_attachment=True,
        download_name=pdf_filename
    )


# INVOICE: Email Invoice PDF (Version 3)
@app.route('/invoices/<int:invoice_id>/email', methods=['POST'])
@login_required
def email_invoice_pdf(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
        
    company = CompanySettings.query.first()
    pdf_filename = f"{invoice.invoice_number}.pdf"
    pdf_path = os.path.join(SECURE_PDF_FOLDER, pdf_filename)
    
    # Regenerate if missing
    if not os.path.exists(pdf_path):
        if not company:
            flash("Cannot send email: Company settings are missing.", "error")
            return redirect(url_for('view_invoice_details', invoice_id=invoice.id))
        try:
            generate_invoice_pdf(invoice, company, pdf_path)
        except Exception as e:
            app.logger.error(f"PDF regeneration failed: {e}")
            flash("Failed to generate PDF attachment file.", "error")
            return redirect(url_for('view_invoice_details', invoice_id=invoice.id))
            
    # Retrieve SMTP configurations
    smtp_server = app.config.get('MAIL_SERVER') or os.environ.get('MAIL_SERVER')
    smtp_port = app.config.get('MAIL_PORT') or os.environ.get('MAIL_PORT') or 587
    smtp_user = app.config.get('MAIL_USERNAME') or os.environ.get('MAIL_USERNAME')
    smtp_pass = app.config.get('MAIL_PASSWORD') or os.environ.get('MAIL_PASSWORD')
    sender_email = app.config.get('MAIL_DEFAULT_SENDER') or smtp_user
    
    if not smtp_server or not smtp_user or not smtp_pass:
        # Step 10: Warn but fall back gracefully if SMTP is not configured
        flash("Email Dispatch Mocked/Not Sent: SMTP Mail Server settings are not configured in environment.", "warning")
        
        # Log mocked dispatch activity
        log = ActivityLog(
            invoice_id=invoice.id,
            invoice_number=invoice.invoice_number,
            action="Email Simulation",
            details=f"Email dispatch simulated to '{invoice.client_email}' (SMTP unconfigured)."
        )
        db.session.add(log)
        db.session.commit()
        return redirect(url_for('view_invoice_details', invoice_id=invoice.id))
        
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        
        # Build MIME message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = invoice.client_email
        msg['Subject'] = f"Invoice {invoice.invoice_number} from {company.name if company else 'BillFlow'}"
        
        body = f"Dear {invoice.client_name},\n\nPlease find attached invoice {invoice.invoice_number} for Rs. {invoice.total_amount:.2f}.\nDue date: {invoice.due_date.strftime('%Y-%m-%d')}.\n\nThank you for your business!\n\nSincerely,\n{company.name if company else 'BillFlow'}"
        msg.attach(MIMEText(body, 'plain'))
        
        # Attachment loading
        with open(pdf_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {pdf_filename}",
        )
        msg.attach(part)
        
        # Connect & Send TLS
        server = smtplib.SMTP(smtp_server, int(smtp_port))
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)
        server.quit()
        
        # Log activity
        log = ActivityLog(
            invoice_id=invoice.id,
            invoice_number=invoice.invoice_number,
            action="Email Sent",
            details=f"Invoice emailed successfully to client at '{invoice.client_email}'."
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f"Invoice {invoice.invoice_number} successfully emailed to {invoice.client_email}!", "success")
        
    except Exception as e:
        app.logger.error(f"Email dispatch error: {e}")
        flash(f"Failed to send email: {str(e)}", "error")
        
    return redirect(url_for('view_invoice_details', invoice_id=invoice.id))


# INVOICE: Mark invoice as Paid
@app.route('/invoices/<int:invoice_id>/mark_paid', methods=['POST'])
@login_required
def mark_paid(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
        
    try:
        invoice.status = 'Paid'
        db.session.commit()
        
        # Regenerate invoice PDF to reflect Paid stamp/status
        company = CompanySettings.query.first()
        pdf_path = os.path.join(SECURE_PDF_FOLDER, f"{invoice.invoice_number}.pdf")
        if company and os.path.exists(pdf_path):
            generate_invoice_pdf(invoice, company, pdf_path)
            
        flash(f"Invoice {invoice.invoice_number} marked as Paid.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Failed to update invoice status.", "error")
        
    return redirect(url_for('dashboard'))


# INVOICE: Delete Invoice
@app.route('/invoices/<int:invoice_id>/delete', methods=['POST'])
@login_required
def delete_invoice(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
        
    invoice_number = invoice.invoice_number
    try:
        # Delete PDF file if exists on disk
        pdf_path = os.path.join(SECURE_PDF_FOLDER, f"{invoice_number}.pdf")
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
            
        # Delete DB row (InvoiceItems delete via CASCADE relation)
        db.session.delete(invoice)
        db.session.commit()
        
        flash(f"Invoice {invoice_number} has been deleted successfully.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Failed to delete invoice: {e}", "error")
        
    return redirect(url_for('dashboard'))


# COMPANY SETTINGS: Edit/setup profile
@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    company = CompanySettings.query.first()
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        gstin = request.form.get('gstin', '').strip().upper()
        
        bank_name = request.form.get('bank_name', '').strip()
        bank_account = request.form.get('bank_account', '').strip()
        bank_ifsc = request.form.get('bank_ifsc', '').strip().upper()
        bank_account_name = request.form.get('bank_account_name', '').strip()
        bank_branch = request.form.get('bank_branch', '').strip()
        upi_id = request.form.get('upi_id', '').strip()
        terms_conditions = request.form.get('terms_conditions', '').strip()
        
        pref_show_hsn_summary = request.form.get('pref_show_hsn_summary') == '1'
        pref_show_bank_details = request.form.get('pref_show_bank_details') == '1'
        pref_show_terms = request.form.get('pref_show_terms') == '1'
        pref_show_notes = request.form.get('pref_show_notes') == '1'
        pref_show_signatory = request.form.get('pref_show_signatory') == '1'
        
        # Validations
        if not name or not email or not address or not phone:
            flash("Company name, email, phone, and address are required.", "error")
            return redirect(url_for('settings'))

        logo_file = request.files.get('logo')
        logo_path = company.logo_path if company else None

        # Process image upload if provided
        if logo_file and logo_file.filename != '':
            if allowed_file(logo_file.filename):
                try:
                    # 1. Inspect image headers using Pillow to assert actual image stream (Magic checks)
                    img = Image.open(logo_file.stream)
                    img.verify() # Throws if image file is invalid/corrupted
                    
                    # 2. Reset stream seek pointer back to start after reading
                    logo_file.stream.seek(0)
                    
                    # 3. Delete previous logo file if exists to clean disk
                    if logo_path and os.path.exists(logo_path):
                        try:
                            os.remove(logo_path)
                        except OSError:
                            pass
                            
                    # 4. Generate collision-free sanitized filename
                    ext = logo_file.filename.rsplit('.', 1)[1].lower()
                    unique_filename = f"logo_{uuid.uuid4().hex}.{ext}"
                    logo_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    
                    # Save image stream
                    logo_file.save(logo_path)
                    
                except Exception as e:
                    flash(f"Invalid logo image: {e}. Please upload a clean image file.", "error")
                    return redirect(url_for('settings'))
            else:
                flash("Unsupported logo file extension. Use PNG, JPG, JPEG, or WEBP.", "error")
                return redirect(url_for('settings'))

        try:
            # Update existing row or create new
            if company:
                company.name = name
                company.email = email
                company.phone = phone
                company.address = address
                company.gstin = gstin if gstin else None
                company.logo_path = logo_path
                company.bank_name = bank_name if bank_name else None
                company.bank_account = bank_account if bank_account else None
                company.bank_ifsc = bank_ifsc if bank_ifsc else None
                company.bank_account_name = bank_account_name if bank_account_name else None
                company.bank_branch = bank_branch if bank_branch else None
                company.upi_id = upi_id if upi_id else None
                company.terms_conditions = terms_conditions if terms_conditions else None
                company.pref_show_hsn_summary = pref_show_hsn_summary
                company.pref_show_bank_details = pref_show_bank_details
                company.pref_show_terms = pref_show_terms
                company.pref_show_notes = pref_show_notes
                company.pref_show_signatory = pref_show_signatory
            else:
                company = CompanySettings(
                    name=name,
                    email=email,
                    phone=phone,
                    address=address,
                    gstin=gstin if gstin else None,
                    logo_path=logo_path,
                    bank_name=bank_name if bank_name else None,
                    bank_account=bank_account if bank_account else None,
                    bank_ifsc=bank_ifsc if bank_ifsc else None,
                    bank_account_name=bank_account_name if bank_account_name else None,
                    bank_branch=bank_branch if bank_branch else None,
                    upi_id=upi_id if upi_id else None,
                    terms_conditions=terms_conditions if terms_conditions else None,
                    pref_show_hsn_summary=pref_show_hsn_summary,
                    pref_show_bank_details=pref_show_bank_details,
                    pref_show_terms=pref_show_terms,
                    pref_show_notes=pref_show_notes,
                    pref_show_signatory=pref_show_signatory
                )
                db.session.add(company)
                
            db.session.commit()
            flash("Company settings updated successfully.", "success")
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Failed to save settings: {e}", "error")
            
    return render_template('settings.html', company=company)


# ==========================================================================
# CUSTOMER DIRECTORY & AUTOMATION ROUTES
# ==========================================================================

# CUSTOMER DIRECTORY: List and Manual Add
@app.route('/customers', methods=['GET', 'POST'])
@login_required
def customers():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        gstin = request.form.get('gstin', '').strip().upper()

        if not name or not email or not address:
            flash("Customer name, email, and address are required.", "error")
            return redirect(url_for('customers'))

        try:
            # Check unique name constraint
            existing = Customer.query.filter_by(name=name).first()
            if existing:
                flash(f"A customer with the name '{name}' already exists.", "error")
                return redirect(url_for('customers'))

            customer = Customer(
                name=name,
                email=email,
                phone=phone if phone else None,
                address=address,
                gstin=gstin if gstin else None
            )
            db.session.add(customer)
            db.session.commit()
            flash(f"Customer '{name}' registered successfully.", "success")
            return redirect(url_for('customers'))
        except Exception as e:
            db.session.rollback()
            flash(f"Failed to add customer: {e}", "error")

    customers = Customer.query.order_by(Customer.name).all()
    return render_template('customers.html', customers=customers)


# CUSTOMER DIRECTORY: Delete Customer
@app.route('/customers/<int:customer_id>/delete', methods=['POST'])
@login_required
def delete_customer(customer_id):
    customer = db.session.get(Customer, customer_id)
    if not customer:
        abort(404)
    try:
        name = customer.name
        db.session.delete(customer)
        db.session.commit()
        flash(f"Customer '{name}' deleted successfully.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Failed to delete customer: {e}", "error")
    return redirect(url_for('customers'))


# CUSTOMER DIRECTORY: Clear Customer Directory (Version 2.8)
@app.route('/customers/clear', methods=['POST'])
@login_required
def clear_customers():
    try:
        db.session.query(Customer).delete()
        db.session.commit()
        flash("Customer directory cleared successfully.", "success")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Failed to clear customer directory: {e}", exc_info=True)
        flash(f"Failed to clear directory: {e}", "error")
    return redirect(url_for('customers'))


# CUSTOMER DIRECTORY: Download Sample Excel Template (Version 2.6)
@app.route('/customers/download-template', methods=['GET'])
@login_required
def download_customer_template():
    try:
        # Create standard Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Template"

        # Headers as specified in guidelines
        headers = ["Company Name", "Email", "Phone", "Address", "GSTIN"]
        ws.append(headers)

        # Style headers to look professional (Indigo primary fill brand color)
        from openpyxl.styles import Font, PatternFill
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

        for col_idx in range(1, 6):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        # Example row
        ws.append(["Acme Corporation", "billing@acme.com", "+91 9876543210", "123 Business Park, Bangalore, KA, 560001", "29AAAAA1111A1Z1"])

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Save workbook to BytesIO RAM buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="billflow_customer_template.xlsx"
        )
    except Exception as e:
        app.logger.error(f"Sample template generation failed: {e}", exc_info=True)
        flash("Failed to generate sample template. Please try again.", "error")
        return redirect(url_for('customers'))


# CUSTOMER DIRECTORY: Excel File Import (Version 2.6 AJAX)
@app.route('/customers/import', methods=['POST'])
@login_required
def import_customers():
    excel_file = request.files.get('excel_file')
    if not excel_file or excel_file.filename == '':
        return {'success': False, 'message': 'Please upload a valid Excel file.'}, 400

    ext = excel_file.filename.rsplit('.', 1)[1].lower() if '.' in excel_file.filename else ''
    if ext not in ['xlsx', 'xls']:
        return {'success': False, 'message': 'Unsupported file format. Please upload an Excel sheet (.xlsx, .xls).'}, 400

    # Security: Process uploads in a temporary location and delete them afterward
    temp_filename = f"import_{uuid.uuid4().hex}.{ext}"
    temp_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
    wb = None
    
    try:
        # Save request file stream to temp file
        excel_file.save(temp_path)
        
        # Load workbook from temp path (Read-only mode safe check)
        wb = openpyxl.load_workbook(temp_path, read_only=True)
        sheet = wb.active

        # Security: Limit imports to 1000 rows of data (row 1 is headers)
        if sheet.max_row > 1001:
            return {'success': False, 'message': 'Import limit exceeded. A spreadsheet can contain a maximum of 1000 rows.'}, 400

        # Read first row for headers
        first_row_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row_cells:
            return {'success': False, 'message': 'Empty spreadsheet uploaded.'}, 400

        # Map header columns dynamically to tolerate variations in casing/spacing
        header_map = {}
        for idx, header in enumerate(first_row_cells):
            if header:
                header_clean = str(header).strip().lower()
                if 'company' in header_clean or 'name' in header_clean:
                    header_map['name'] = idx
                elif 'email' in header_clean:
                    header_map['email'] = idx
                elif 'phone' in header_clean:
                    header_map['phone'] = idx
                elif 'address' in header_clean:
                    header_map['address'] = idx
                elif 'gstin' in header_clean:
                    header_map['gstin'] = idx

        # Security: Validate headers before processing
        if 'name' not in header_map or 'email' not in header_map or 'address' not in header_map:
            return {'success': False, 'message': 'Required columns are missing. Please download and use the sample template.'}, 400

        imported_count = 0
        duplicate_count = 0
        invalid_count = 0
        staged_names = set()
        staged_gstins = set()
        report = []
        import re

        email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'

        # Read rows starting from row 2
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if not any(cell is not None and str(cell).strip() != '' for cell in row):
                continue
                
            # Validation: Trim unnecessary whitespace
            name = str(row[header_map['name']]).strip() if header_map['name'] < len(row) and row[header_map['name']] is not None else ''
            email = str(row[header_map['email']]).strip() if header_map['email'] < len(row) and row[header_map['email']] is not None else ''
            address = str(row[header_map['address']]).strip() if header_map['address'] < len(row) and row[header_map['address']] is not None else ''
            
            # Validation: Skip invalid rows and log reason
            if not name:
                invalid_count += 1
                report.append({'row': row_idx, 'status': 'Skipped', 'reason': 'Name is missing.'})
                continue
            if not email:
                invalid_count += 1
                report.append({'row': row_idx, 'status': 'Skipped', 'reason': 'Email is missing.'})
                continue
            if not address:
                invalid_count += 1
                report.append({'row': row_idx, 'status': 'Skipped', 'reason': 'Billing Address is missing.'})
                continue

            # Validate Email syntax format
            if not re.match(email_regex, email):
                invalid_count += 1
                report.append({'row': row_idx, 'status': 'Skipped', 'reason': f"Invalid email format: '{email}'"})
                continue

            phone = str(row[header_map['phone']]).strip() if ('phone' in header_map and header_map['phone'] < len(row) and row[header_map['phone']] is not None) else None
            gstin = str(row[header_map['gstin']]).strip().upper() if ('gstin' in header_map and header_map['gstin'] < len(row) and row[header_map['gstin']] is not None) else None

            # Validation: Prevent duplicate customer entries (Primary: GSTIN, Secondary: Company Name)
            is_duplicate = False
            skip_reason = ""
            
            if gstin:
                if gstin.lower() in staged_gstins:
                    is_duplicate = True
                    skip_reason = f"Duplicate GSTIN in spreadsheet: '{gstin}'"
                else:
                    existing_gstin = Customer.query.filter(Customer.gstin.ilike(gstin)).first()
                    if existing_gstin:
                        is_duplicate = True
                        skip_reason = f"Customer with GSTIN '{gstin}' already exists in Directory (Company: '{existing_gstin.name}')."
            
            if not is_duplicate:
                if name.lower() in staged_names:
                    is_duplicate = True
                    skip_reason = f"Duplicate Company Name in spreadsheet: '{name}'"
                else:
                    existing_name = Customer.query.filter_by(name=name).first()
                    if existing_name:
                        is_duplicate = True
                        skip_reason = f"Customer with Company Name '{name}' already exists in Directory."
            
            if is_duplicate:
                duplicate_count += 1
                report.append({'row': row_idx, 'status': 'Skipped', 'reason': skip_reason})
                continue

            customer = Customer(
                name=name,
                email=email,
                phone=phone if phone else None,
                address=address,
                gstin=gstin if gstin else None
            )
            db.session.add(customer)
            staged_names.add(name.lower())
            if gstin:
                staged_gstins.add(gstin.lower())
            imported_count += 1

        # Save all valid rows in a single database transaction with rollback on failure
        db.session.commit()
        
        return {
            'success': True,
            'imported_count': imported_count,
            'duplicate_count': duplicate_count,
            'invalid_count': invalid_count,
            'report': report
        }
        
    except Exception as e:
        db.session.rollback()
        # Log full traceback without exposing details to client
        app.logger.error(f"Customer Excel Import Failed: {e}", exc_info=True)
        return {
            'success': False,
            'message': 'The uploaded file is corrupted or cannot be read.'
        }, 500
        
    finally:
        # Close openpyxl workbook handle to release file locks on Windows
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
                
        # Security: Clean up the temp file from disk in a finally block
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError as e:
                app.logger.error(f"Failed to delete temp file {temp_path}: {e}")


# CUSTOMER DIRECTORY: Bulk JSON Create (Version 2.5)
@app.route('/customers/bulk-create', methods=['POST'])
@login_required
def bulk_create_customers():
    data = request.get_json()
    if not data or 'rows' not in data:
        return {'success': False, 'message': 'Invalid payload.'}, 400

    rows = data['rows']
    
    # Security: Limit imports to 1000 rows
    if len(rows) > 1000:
        return {'success': False, 'message': 'Spreadsheet limit exceeded. Maximum allowed rows is 1000.'}, 400

    added_count = 0
    duplicate_count = 0
    error_rows = []
    staged_names = set()
    staged_gstins = set()

    try:
        # Run validations inside database transaction context
        for idx, row in enumerate(rows):
            name = row.get('name', '').strip()
            email = row.get('email', '').strip()
            phone = row.get('phone', '').strip()
            address = row.get('address', '').strip()
            gstin = row.get('gstin', '').strip().upper()

            # Ignore completely empty rows
            if not name and not email and not phone and not address and not gstin:
                continue

            # Validate each row independently
            row_errors = {}
            if not name:
                row_errors['name'] = 'Customer Name is required.'
            if not email:
                row_errors['email'] = 'Email Address is required.'
            if not address:
                row_errors['address'] = 'Billing Address is required.'

            if row_errors:
                error_rows.append({
                    'row_idx': idx,
                    'errors': row_errors
                })
                continue

            # Validation: Prevent duplicate customer entries (Primary: GSTIN, Secondary: Company Name)
            is_duplicate = False
            
            if gstin:
                if gstin.lower() in staged_gstins:
                    is_duplicate = True
                else:
                    existing_gstin = Customer.query.filter(Customer.gstin.ilike(gstin)).first()
                    if existing_gstin:
                        is_duplicate = True
            
            if not is_duplicate:
                if name.lower() in staged_names:
                    is_duplicate = True
                else:
                    existing_name = Customer.query.filter_by(name=name).first()
                    if existing_name:
                        is_duplicate = True
            
            if is_duplicate:
                duplicate_count += 1
                continue

            customer = Customer(
                name=name,
                email=email,
                phone=phone if phone else None,
                address=address,
                gstin=gstin if gstin else None
            )
            db.session.add(customer)
            staged_names.add(name.lower())
            if gstin:
                staged_gstins.add(gstin.lower())
            added_count += 1

        # Commit only valid additions in a single transaction with rollback on failure
        if added_count > 0:
            db.session.commit()
        else:
            db.session.rollback()

        # Build clean confirmation summary
        err_msg = ""
        if error_rows:
            err_msg = f" {len(error_rows)} row(s) contain validation errors."
        
        status_msg = f"{added_count} customers added successfully.{' ' + str(duplicate_count) + ' duplicate customers skipped.' if duplicate_count > 0 else ''}{err_msg}"

        return {
            'success': True,
            'added_count': added_count,
            'duplicate_count': duplicate_count,
            'error_rows': error_rows,
            'message': status_msg
        }

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Bulk customer registration failed: {e}")
        return {
            'success': False,
            'message': f"A database error occurred: {str(e)}"
        }, 500


# CUSTOMER DIRECTORY: JSON API for Auto-fill Autocomplete
@app.route('/api/customers/<int:customer_id>')
@login_required
def get_customer_json(customer_id):
    customer = db.session.get(Customer, customer_id)
    if not customer:
        return {'error': 'Customer not found'}, 404
        
    return {
        'name': customer.name,
        'email': customer.email,
        'phone': customer.phone if customer.phone else '',
        'address': customer.address,
        'gstin': customer.gstin if customer.gstin else ''
    }


# ==========================================================================
# INVOICE VERSION 2 ROUTING CONTROLLERS (EDIT / DUPLICATE)
# ==========================================================================

# INVOICE: Edit invoice route
@app.route('/invoices/<int:invoice_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_invoice(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
        
    company = CompanySettings.query.first()
    if not company:
        flash("You must complete company settings before modifying invoices.", "warning")
        return redirect(url_for('settings'))

    if request.method == 'POST':
        client_name = request.form.get('client_name', '').strip()
        client_email = request.form.get('client_email', '').strip()
        client_phone = request.form.get('client_phone', '').strip()
        client_address = request.form.get('client_address', '').strip()
        client_gstin = request.form.get('client_gstin', '').strip().upper()
        
        status = request.form.get('status', 'Unpaid')
        notes = request.form.get('notes', '').strip()
        
        # Validations
        if not client_name or not client_email or not client_address:
            flash("Client name, email, and address are required.", "error")
            return redirect(url_for('edit_invoice', invoice_id=invoice.id))
            
        try:
            gst_rate = float(request.form.get('gst_rate', 18.0))
            discount = float(request.form.get('discount', 0.0))
            if discount < 0 or discount > 100:
                discount = 0.0
        except ValueError:
            gst_rate = 18.0
            discount = 0.0

        # Retrieve arrays
        descriptions = request.form.getlist('description[]')
        hsn_sacs = request.form.getlist('hsn_sac[]')
        quantities = request.form.getlist('quantity[]')
        unit_prices = request.form.getlist('unit_price[]')
        
        if not descriptions or len(descriptions) == 0:
            flash("An invoice must contain at least one line item.", "error")
            return redirect(url_for('edit_invoice', invoice_id=invoice.id))

        subtotal = 0.0
        new_items = []
        
        try:
            for i in range(len(descriptions)):
                desc = descriptions[i].strip()
                hsn = hsn_sacs[i].strip() if (hsn_sacs and i < len(hsn_sacs) and hsn_sacs[i].strip()) else None
                qty = int(quantities[i])
                price = float(unit_prices[i])
                
                # Check negative constraints
                if qty < 1 or price < 0:
                    raise ValueError("Quantity must be >= 1 and Unit Price >= 0.")
                    
                item_total = qty * price
                subtotal += item_total
                
                item = InvoiceItem(
                    description=desc,
                    quantity=qty,
                    unit_price=price,
                    tax_rate=gst_rate,
                    hsn_sac=hsn,
                    total=item_total
                )
                new_items.append(item)
        except (ValueError, IndexError) as e:
            flash(f"Invalid item data: {str(e)}", "error")
            return redirect(url_for('edit_invoice', invoice_id=invoice.id))

        # Recompute totals on backend
        discount_amount = subtotal * (discount / 100)
        taxable_amount = subtotal - discount_amount
        tax_amount = taxable_amount * (gst_rate / 100)
        total_amount = taxable_amount + tax_amount

        try:
            # Parse dates
            date_created_str = request.form.get('date_created')
            due_date_str = request.form.get('due_date')
            
            if date_created_str:
                invoice.date_created = datetime.strptime(date_created_str, '%Y-%m-%d').date()
            if due_date_str:
                invoice.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()

            # Calculate CGST, SGST, IGST based on State Codes (first 2 digits of GSTIN)
            company = CompanySettings.query.first()
            company_state = company.gstin[:2] if (company and company.gstin and len(company.gstin) >= 2) else None
            client_state = client_gstin[:2] if (client_gstin and len(client_gstin) >= 2) else None
            
            if company_state and client_state and company_state == client_state:
                cgst = tax_amount / 2.0
                sgst = tax_amount / 2.0
                igst = 0.0
            else:
                cgst = 0.0
                sgst = 0.0
                igst = tax_amount

            # Overwrite fields
            invoice.status = status
            invoice.client_name = client_name
            invoice.client_email = client_email
            invoice.client_phone = client_phone if client_phone else None
            invoice.client_address = client_address
            invoice.client_gstin = client_gstin if client_gstin else None
            invoice.gst_rate = gst_rate
            invoice.discount = discount
            invoice.subtotal = subtotal
            invoice.tax_amount = tax_amount
            invoice.cgst = cgst
            invoice.sgst = sgst
            invoice.igst = igst
            invoice.total_amount = total_amount
            invoice.notes = notes if notes else None

            # Remove old items from DB relationship, add new ones
            invoice.items.clear() # database triggers cascade deletion of orphaned InvoiceItems
            for item in new_items:
                invoice.items.append(item)
                
            db.session.commit()

            # Compile updated PDF
            pdf_filename = f"{invoice.invoice_number}.pdf"
            pdf_path = os.path.join(SECURE_PDF_FOLDER, pdf_filename)
            generate_invoice_pdf(invoice, company, pdf_path)
            
            flash(f"Invoice {invoice.invoice_number} updated successfully!", "success")
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Invoice modification failed: {e}")
            flash("Failed to save changes. Please try again.", "error")
            return redirect(url_for('edit_invoice', invoice_id=invoice.id))

    customers = Customer.query.order_by(Customer.name).all()
    return render_template('edit_invoice.html', invoice=invoice, company=company, customers=customers)


# INVOICE: Duplicate invoice route
@app.route('/invoices/<int:invoice_id>/duplicate')
@login_required
def duplicate_invoice(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        abort(404)
        
    company = CompanySettings.query.first()
    if not company:
        flash("You must complete company settings first.", "warning")
        return redirect(url_for('settings'))

    try:
        # Generate new serial invoice code
        new_invoice_number = get_next_invoice_number()
        today = datetime.now(timezone.utc).date()
        due = today + timedelta(days=15)

        # Build duplicated instance
        duplicated_invoice = Invoice(
            invoice_number=new_invoice_number,
            date_created=today,
            due_date=due,
            status='Unpaid',
            client_name=invoice.client_name,
            client_email=invoice.client_email,
            client_phone=invoice.client_phone,
            client_address=invoice.client_address,
            client_gstin=invoice.client_gstin,
            gst_rate=invoice.gst_rate,
            discount=invoice.discount,
            subtotal=invoice.subtotal,
            tax_amount=invoice.tax_amount,
            cgst=invoice.cgst,
            sgst=invoice.sgst,
            igst=invoice.igst,
            total_amount=invoice.total_amount,
            notes=invoice.notes
        )

        # Clone and link items
        for item in invoice.items:
            cloned_item = InvoiceItem(
                description=item.description,
                quantity=item.quantity,
                unit_price=item.unit_price,
                tax_rate=item.tax_rate,
                hsn_sac=item.hsn_sac,
                total=item.total
            )
            duplicated_invoice.items.append(cloned_item)

        db.session.add(duplicated_invoice)
        db.session.commit()

        # Compile PDF file
        pdf_path = os.path.join(SECURE_PDF_FOLDER, f"{new_invoice_number}.pdf")
        generate_invoice_pdf(duplicated_invoice, company, pdf_path)

        flash(f"Invoice {invoice.invoice_number} duplicated successfully as {new_invoice_number}!", "success")
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Invoice duplication failed: {e}")
        flash("Failed to duplicate invoice.", "error")
        
    return redirect(url_for('dashboard'))


# ==========================================================================
# REPORTS: Revenue Comparison (Version 4 Phase 5)
# ==========================================================================

@app.route('/api/reports/revenue-comparison')
@login_required
def api_reports_revenue_comparison():
    """
    Compares revenue between current vs previous month,
    and current vs previous year.
    Returns revenue totals, invoice counts, differences and % growth.
    """
    try:
        today = get_ist_now().date()

        # ---- Month Comparison ----
        curr_month_start = today.replace(day=1)
        # Previous month: go back to last day of previous month, then to 1st
        prev_month_end = curr_month_start - timedelta(days=1)
        prev_month_start = prev_month_end.replace(day=1)

        # ---- Year Comparison ----
        curr_year_start = today.replace(month=1, day=1)
        prev_year_start = curr_year_start.replace(year=today.year - 1)
        prev_year_end = curr_year_start - timedelta(days=1)

        def fetch_period_stats(start, end):
            row = db.session.query(
                db.func.coalesce(db.func.sum(db.case(
                    (Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue']), Invoice.total_amount),
                    else_=0.0
                )), 0.0).label('revenue'),
                db.func.count(Invoice.id).label('invoice_count')
            ).filter(
                Invoice.is_deleted == False,
                Invoice.date_created.between(start, end)
            ).first()
            return float(row.revenue or 0.0), int(row.invoice_count or 0)

        curr_month_rev, curr_month_count = fetch_period_stats(curr_month_start, today)
        prev_month_rev, prev_month_count = fetch_period_stats(prev_month_start, prev_month_end)

        curr_year_rev, curr_year_count = fetch_period_stats(curr_year_start, today)
        prev_year_rev, prev_year_count = fetch_period_stats(prev_year_start, prev_year_end)

        def calc_growth(current, previous):
            if previous > 0:
                return round(((current - previous) / previous) * 100, 2)
            elif current > 0:
                return 100.0  # New revenue where none existed
            return 0.0

        return jsonify({
            'success': True,
            'month': {
                'current_revenue': curr_month_rev,
                'previous_revenue': prev_month_rev,
                'revenue_diff': curr_month_rev - prev_month_rev,
                'growth_pct': calc_growth(curr_month_rev, prev_month_rev),
                'current_count': curr_month_count,
                'previous_count': prev_month_count,
                'count_diff': curr_month_count - prev_month_count,
                'current_label': curr_month_start.strftime('%B %Y'),
                'previous_label': prev_month_start.strftime('%B %Y')
            },
            'year': {
                'current_revenue': curr_year_rev,
                'previous_revenue': prev_year_rev,
                'revenue_diff': curr_year_rev - prev_year_rev,
                'growth_pct': calc_growth(curr_year_rev, prev_year_rev),
                'current_count': curr_year_count,
                'previous_count': prev_year_count,
                'count_diff': curr_year_count - prev_year_count,
                'current_label': str(today.year),
                'previous_label': str(today.year - 1)
            }
        })

    except Exception as e:
        app.logger.error(f"Revenue comparison failed: {e}", exc_info=True)
        return jsonify({'success': False, 'message': 'Failed to load revenue comparison data.'}), 500


# ==========================================================================
# REPORTS: Shared Export Helper Utilities (Version 4 Export Module)
# ==========================================================================

def _style_excel_header(ws, headers, fill_color='4F46E5'):
    """Apply BillFlow branded header styles to an openpyxl worksheet row 1."""
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 18


def _auto_col_width(ws):
    """Auto-fit column widths based on cell content."""
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def _build_report_pdf(title, headers, rows, totals_row=None):
    """
    Generate a styled PDF report using ReportLab.
    Returns BytesIO buffer containing the PDF data.

    Args:
        title: Report title string
        headers: List of column header strings
        rows: List of lists (each row is a list of cell values)
        totals_row: Optional list of values for a totals/summary row
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    buf = BytesIO()
    page_size = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    brand_color = colors.HexColor('#4F46E5')
    dark_color = colors.HexColor('#1e293b')
    light_gray = colors.HexColor('#f8fafc')
    border_gray = colors.HexColor('#e2e8f0')
    green_color = colors.HexColor('#10B981')

    title_style = ParagraphStyle(
        'Title', parent=styles['Title'],
        fontSize=18, textColor=dark_color, spaceAfter=4,
        fontName='Helvetica-Bold'
    )
    sub_style = ParagraphStyle(
        'Sub', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#64748b'), spaceAfter=12
    )

    story = []
    story.append(Paragraph('BillFlow', ParagraphStyle(
        'Brand', parent=styles['Normal'],
        fontSize=10, textColor=brand_color, fontName='Helvetica-Bold', spaceAfter=2
    )))
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(
        f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
        sub_style
    ))
    story.append(Spacer(1, 0.3*cm))

    # Table data: header + rows + optional totals
    table_data = [headers] + rows
    if totals_row:
        table_data.append(totals_row)

    # Build column widths
    available_width = page_size[0] - 3*cm
    col_count = len(headers)
    col_width = available_width / col_count

    tbl = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)

    table_styles = [
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), brand_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ROWHEIGHT', (0, 0), (-1, 0), 18),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Data rows – alternating backgrounds
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWHEIGHT', (0, 1), (-1, -1), 15),
        ('GRID', (0, 0), (-1, -1), 0.3, border_gray),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]

    # Alternating row shading
    for row_idx in range(1, len(table_data)):
        if totals_row and row_idx == len(table_data) - 1:
            # Totals row
            table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#EEF2FF')))
            table_styles.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
        elif row_idx % 2 == 0:
            table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), light_gray))
        else:
            table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.white))

    tbl.setStyle(TableStyle(table_styles))
    story.append(tbl)

    if totals_row:
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            f'Total records: {len(rows)}',
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#64748b'))
        ))

    doc.build(story)
    buf.seek(0)
    return buf


def _get_export_params():
    """Parse and validate export parameters from request args."""
    fmt = request.args.get('format', 'excel').strip().lower()
    if fmt not in ('excel', 'csv', 'pdf'):
        fmt = 'excel'
    try:
        start_date, end_date = parse_report_dates()
    except ValueError as e:
        raise ValueError(str(e))
    return fmt, start_date, end_date


# ==========================================================================
# REPORTS: Export – Revenue Report (Version 4)
# ==========================================================================

@app.route('/api/reports/export/revenue')
@login_required
def export_report_revenue():
    """Export revenue trend data as Excel, CSV, or PDF."""
    try:
        fmt, start_date, end_date = _get_export_params()
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400

    # Fetch revenue data (same query as /api/reports/revenue)
    results = db.session.query(
        Invoice.date_created,
        db.func.count(Invoice.id).label('invoice_count'),
        db.func.sum(Invoice.total_amount).label('total_revenue'),
        db.func.sum(Invoice.tax_amount).label('total_tax'),
        db.func.sum(Invoice.subtotal).label('total_subtotal')
    ).filter(
        Invoice.is_deleted == False,
        Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue']),
        Invoice.date_created.between(start_date, end_date)
    ).group_by(Invoice.date_created).order_by(Invoice.date_created.asc()).all()

    headers = ['Date', 'Invoice Count', 'Subtotal (₹)', 'Tax Amount (₹)', 'Total Revenue (₹)']
    rows = []
    grand_total = 0.0
    for r in results:
        rev = float(r.total_revenue or 0.0)
        grand_total += rev
        rows.append([
            r.date_created.strftime('%Y-%m-%d'),
            r.invoice_count or 0,
            f"{float(r.total_subtotal or 0.0):.2f}",
            f"{float(r.total_tax or 0.0):.2f}",
            f"{rev:.2f}"
        ])

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    title = f'Revenue Report | {start_date} to {end_date}'

    if fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Revenue Report'
        _style_excel_header(ws, headers)
        for row in rows:
            ws.append(row)
        # Totals row
        ws.append(['TOTAL', sum(r[1] for r in results), '', '', f"{grand_total:.2f}"])
        bold_font = Font(bold=True)
        for cell in ws[ws.max_row]:
            cell.font = bold_font
        _auto_col_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'revenue_report_{timestamp}.xlsx')

    elif fmt == 'csv':
        buf = BytesIO()
        buf.write('\ufeff'.encode('utf-8'))  # BOM for Excel compatibility
        import io as _io
        text_buf = _io.StringIO()
        writer = csv.writer(text_buf)
        writer.writerow(headers)
        writer.writerows(rows)
        writer.writerow(['TOTAL', sum(r[1] for r in results), '', '', f"{grand_total:.2f}"])
        buf.write(text_buf.getvalue().encode('utf-8'))
        buf.seek(0)
        return send_file(buf, mimetype='text/csv', as_attachment=True,
                         download_name=f'revenue_report_{timestamp}.csv')

    else:  # pdf
        totals_row = ['TOTAL', str(sum(r[1] for r in results)), '', '', f"{grand_total:.2f}"]
        pdf_buf = _build_report_pdf(title, headers, rows, totals_row)
        return send_file(pdf_buf, mimetype='application/pdf', as_attachment=True,
                         download_name=f'revenue_report_{timestamp}.pdf')


# ==========================================================================
# REPORTS: Export – Customer Report (Version 4)
# ==========================================================================

@app.route('/api/reports/export/customers')
@login_required
def export_report_customers():
    """Export customer analytics as Excel, CSV, or PDF."""
    try:
        fmt, start_date, end_date = _get_export_params()
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400

    # Same query as /api/reports/customers
    results = db.session.query(
        Customer.name,
        Customer.email,
        db.func.coalesce(db.func.sum(db.case(
            (Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue']), Invoice.total_amount),
            else_=0.0
        )), 0.0).label('revenue'),
        db.func.count(Invoice.id).label('invoice_count'),
        db.func.coalesce(db.func.sum(db.case(
            (Invoice.status.in_(['Sent', 'Pending', 'Overdue', 'Unpaid']), Invoice.total_amount),
            else_=0.0
        )), 0.0).label('outstanding'),
        db.func.coalesce(db.func.avg(Invoice.total_amount), 0.0).label('avg_value'),
        db.func.max(Invoice.date_created).label('last_date')
    ).select_from(Customer)\
     .outerjoin(Invoice, (Invoice.customer_id == Customer.id) & (Invoice.is_deleted == False) & (Invoice.date_created.between(start_date, end_date)))\
     .group_by(Customer.id, Customer.name, Customer.email)\
     .order_by(db.desc('revenue')).all()

    headers = ['Customer Name', 'Email', 'Total Revenue (₹)', 'Invoice Count', 'Avg Invoice Value (₹)', 'Outstanding (₹)', 'Last Invoice Date']
    rows = []
    total_rev = 0.0
    total_outstanding = 0.0
    for r in results:
        rev = float(r.revenue or 0.0)
        outs = float(r.outstanding or 0.0)
        total_rev += rev
        total_outstanding += outs
        rows.append([
            r.name,
            r.email,
            f"{rev:.2f}",
            r.invoice_count or 0,
            f"{float(r.avg_value or 0.0):.2f}",
            f"{outs:.2f}",
            r.last_date.strftime('%Y-%m-%d') if r.last_date else '-'
        ])

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    title = f'Customer Analytics Report | {start_date} to {end_date}'

    if fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Customer Report'
        _style_excel_header(ws, headers)
        for row in rows:
            ws.append(row)
        ws.append(['TOTAL', '', f"{total_rev:.2f}", '', '', f"{total_outstanding:.2f}", ''])
        bold_font = Font(bold=True)
        for cell in ws[ws.max_row]:
            cell.font = bold_font
        _auto_col_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'customer_report_{timestamp}.xlsx')

    elif fmt == 'csv':
        buf = BytesIO()
        import io as _io
        text_buf = _io.StringIO()
        writer = csv.writer(text_buf)
        writer.writerow(headers)
        writer.writerows(rows)
        writer.writerow(['TOTAL', '', f"{total_rev:.2f}", '', '', f"{total_outstanding:.2f}", ''])
        buf.write('\ufeff'.encode('utf-8'))
        buf.write(text_buf.getvalue().encode('utf-8'))
        buf.seek(0)
        return send_file(buf, mimetype='text/csv', as_attachment=True,
                         download_name=f'customer_report_{timestamp}.csv')

    else:  # pdf
        totals_row = ['TOTAL', '', f"{total_rev:.2f}", '', '', f"{total_outstanding:.2f}", '']
        pdf_buf = _build_report_pdf(title, headers, rows, totals_row)
        return send_file(pdf_buf, mimetype='application/pdf', as_attachment=True,
                         download_name=f'customer_report_{timestamp}.pdf')


# ==========================================================================
# REPORTS: Export – Invoice Report (Version 4)
# ==========================================================================

@app.route('/api/reports/export/invoices')
@login_required
def export_report_invoices():
    """Export full invoice list as Excel, CSV, or PDF."""
    try:
        fmt, start_date, end_date = _get_export_params()
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400

    invoices = Invoice.query.filter(
        Invoice.is_deleted == False,
        Invoice.date_created.between(start_date, end_date)
    ).order_by(Invoice.date_created.desc()).all()

    headers = ['Invoice No', 'Date', 'Due Date', 'Client Name', 'Client GSTIN', 'Subtotal (₹)', 'Discount (%)', 'Tax (₹)', 'Total (₹)', 'Status']
    rows = []
    grand_total = 0.0
    for inv in invoices:
        grand_total += inv.total_amount
        rows.append([
            inv.invoice_number,
            inv.date_created.strftime('%Y-%m-%d'),
            inv.due_date.strftime('%Y-%m-%d'),
            inv.client_name,
            inv.client_gstin or '-',
            f"{inv.subtotal:.2f}",
            f"{inv.discount:.1f}%",
            f"{inv.tax_amount:.2f}",
            f"{inv.total_amount:.2f}",
            inv.status
        ])

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    title = f'Invoice Report | {start_date} to {end_date}'

    if fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Invoice Report'
        _style_excel_header(ws, headers)

        status_colors = {
            'Paid': 'D1FAE5', 'Overdue': 'FEE2E2', 'Cancelled': 'F3F4F6',
            'Pending': 'DBEAFE', 'Sent': 'DBEAFE', 'Draft': 'FFF3CD', 'Unpaid': 'FFF3CD'
        }
        for row_idx, row in enumerate(rows, start=2):
            ws.append(row)
            status = row[9]
            fill_hex = status_colors.get(status, 'FFFFFF')
            status_fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
            ws.cell(row=row_idx, column=10).fill = status_fill

        ws.append(['TOTAL', '', '', '', '', '', '', '', f"{grand_total:.2f}", ''])
        bold_font = Font(bold=True)
        for cell in ws[ws.max_row]:
            cell.font = bold_font
        _auto_col_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'invoice_report_{timestamp}.xlsx')

    elif fmt == 'csv':
        buf = BytesIO()
        import io as _io
        text_buf = _io.StringIO()
        writer = csv.writer(text_buf)
        writer.writerow(headers)
        writer.writerows(rows)
        writer.writerow(['TOTAL', '', '', '', '', '', '', '', f"{grand_total:.2f}", ''])
        buf.write('\ufeff'.encode('utf-8'))
        buf.write(text_buf.getvalue().encode('utf-8'))
        buf.seek(0)
        return send_file(buf, mimetype='text/csv', as_attachment=True,
                         download_name=f'invoice_report_{timestamp}.csv')

    else:  # pdf
        totals_row = ['TOTAL', '', '', '', '', '', '', '', f"{grand_total:.2f}", '']
        pdf_buf = _build_report_pdf(title, headers, rows, totals_row)
        return send_file(pdf_buf, mimetype='application/pdf', as_attachment=True,
                         download_name=f'invoice_report_{timestamp}.pdf')


# ==========================================================================
# REPORTS: Export – GST Report (Version 4)
# ==========================================================================

@app.route('/api/reports/export/gst')
@login_required
def export_report_gst():
    """Export GST ledger as Excel, CSV, or PDF."""
    try:
        fmt, start_date, end_date = _get_export_params()
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400

    # Same query as /api/reports/gst
    invoices = Invoice.query.filter(
        Invoice.is_deleted == False,
        Invoice.status.in_(['Paid', 'Sent', 'Pending', 'Overdue']),
        Invoice.date_created.between(start_date, end_date)
    ).order_by(Invoice.date_created.asc()).all()

    headers = ['Invoice No', 'Date', 'Client Name', 'Client GSTIN', 'Taxable Value (₹)', 'CGST (₹)', 'SGST (₹)', 'IGST (₹)', 'Total GST (₹)', 'Grand Total (₹)']
    rows = []
    totals = {'taxable': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0, 'gst': 0.0, 'grand': 0.0}

    for inv in invoices:
        discount_amount = inv.subtotal * (inv.discount / 100.0)
        taxable_value = inv.subtotal - discount_amount
        totals['taxable'] += taxable_value
        totals['cgst'] += float(inv.cgst or 0.0)
        totals['sgst'] += float(inv.sgst or 0.0)
        totals['igst'] += float(inv.igst or 0.0)
        totals['gst'] += float(inv.tax_amount or 0.0)
        totals['grand'] += float(inv.total_amount or 0.0)

        rows.append([
            inv.invoice_number,
            inv.date_created.strftime('%Y-%m-%d'),
            inv.client_name,
            inv.client_gstin or '-',
            f"{taxable_value:.2f}",
            f"{float(inv.cgst or 0.0):.2f}",
            f"{float(inv.sgst or 0.0):.2f}",
            f"{float(inv.igst or 0.0):.2f}",
            f"{float(inv.tax_amount or 0.0):.2f}",
            f"{float(inv.total_amount or 0.0):.2f}"
        ])

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    title = f'GST Report | {start_date} to {end_date}'
    totals_row = ['TOTAL', '', '', '',
                  f"{totals['taxable']:.2f}",
                  f"{totals['cgst']:.2f}",
                  f"{totals['sgst']:.2f}",
                  f"{totals['igst']:.2f}",
                  f"{totals['gst']:.2f}",
                  f"{totals['grand']:.2f}"]

    if fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'GST Report'
        _style_excel_header(ws, headers, fill_color='059669')  # Green for GST
        for row in rows:
            ws.append(row)
        ws.append(totals_row)
        bold_font = Font(bold=True)
        for cell in ws[ws.max_row]:
            cell.font = bold_font
        _auto_col_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'gst_report_{timestamp}.xlsx')

    elif fmt == 'csv':
        buf = BytesIO()
        import io as _io
        text_buf = _io.StringIO()
        writer = csv.writer(text_buf)
        writer.writerow(headers)
        writer.writerows(rows)
        writer.writerow(totals_row)
        buf.write('\ufeff'.encode('utf-8'))
        buf.write(text_buf.getvalue().encode('utf-8'))
        buf.seek(0)
        return send_file(buf, mimetype='text/csv', as_attachment=True,
                         download_name=f'gst_report_{timestamp}.csv')

    else:  # pdf
        pdf_buf = _build_report_pdf(title, headers, rows, totals_row)
        return send_file(pdf_buf, mimetype='application/pdf', as_attachment=True,
                         download_name=f'gst_report_{timestamp}.pdf')


# ==========================================================================
# 5. Application Factory Self-Bootstrapping Engine
# ==========================================================================
if __name__ == '__main__':
    # Boot tables & default seeds inside application context
    with app.app_context():
        # Create database file and tables if missing
        db.create_all()
        
        # Seed default administrator credentials if table is empty
        if User.query.count() == 0:
            # Define default login details
            admin_user = User(
                full_name='Administrator',
                company_name='BillFlow Inc',
                email='admin@billflow.com',
                password_hash=generate_password_hash('AdminPassword123')
            )
            db.session.add(admin_user)
            db.session.commit()
            
            print("\n" + "="*60)
            print("DATABASE SEED: Default admin user created successfully!")
            print("  Email: admin@billflow.com")
            print("  Password: AdminPassword123")
            print("  (Please use these credentials to log in to the portal)")
            print("="*60 + "\n")
            
    # Run dev server locally on port 5000
    debug_mode = os.environ.get('FLASK_ENV', 'development') != 'production'
    app.run(debug=debug_mode, port=5000)
